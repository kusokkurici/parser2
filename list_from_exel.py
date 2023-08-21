import openpyxl


def read_website_column(file_path, sheet_name, column_name):
    """Функция считывает данные из файла эксель, возвращает список данных (адреса сайтов) из заданного столбика"""
    website_list = []

    # Открываем файл Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Находим номер столбца "Веб-сайт 1" по его названию
    column_index = None
    for col_num, cell in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column, values_only=True)):
        if cell[0] == column_name:
            column_index = col_num + 1
            break

    if column_index is None:
        return print(f"Столбец '{column_name}' не найден в файле.")

    # Считываем значения из столбца "Веб-сайт 1" в список, пропуская пустые строки
    for row_num, cell_value in enumerate(sheet.iter_cols(min_col=column_index, max_col=column_index, values_only=True)):
        list_url = list(cell_value)
        for i in list_url[1:]:
            if i is not None:
                website_list.append(i)
    return website_list
